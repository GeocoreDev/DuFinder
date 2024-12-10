from PIL import Image
import pytesseract
from PIL.ExifTags import TAGS
import PyPDF2
from datetime import datetime
import comtypes.client
import openpyxl
from docx import Document
import win32com.client
import os
import textract
import re
import threading
import regex  # Asegúrate de tener la biblioteca regex instalada
import olefile  # Asegúrate de tener la biblioteca olefile instalada

# ------    Tools  -------
def parse_date(date_string):
    try:
        if isinstance(date_string, datetime):
            # Si la entrada es del tipo datetime.datetime, devolverla directamente
            return date_string.strftime('%Y-%m-%d %H:%M:%S')

        # Eliminar 'D:' si está presente en el formato
        date_string = date_string.replace('D:', '', 1) if date_string.startswith('D:') else date_string

        # Convertir la cadena de fecha a un objeto datetime
        parsed_date = datetime.strptime(date_string, '%Y%m%d%H%M%S')

        # Formatear la fecha en un formato legible
        formatted_date = parsed_date.strftime('%Y-%m-%d %H:%M:%S')

        return formatted_date
    except ValueError as e:
        return f'Error de formato: {e}'


def convert_gps_coordinates(gps_info):
    gps_coordinates = {}

    for key in gps_info:
        if key in [2, 4]:
            # Obteniendo los grados, minutos y segundos
            degrees, minutes, seconds = gps_info[key]

            # Convirtiendo a grados decimales
            decimal_degrees = degrees + (minutes / 60.0) + (seconds / 3600.0)

            # Agregando al diccionario de coordenadas GPS
            if key == 2:
                gps_coordinates['Latitude'] = decimal_degrees if gps_info[1] == 'N' else -decimal_degrees
            elif key == 4:
                gps_coordinates['Longitude'] = decimal_degrees if gps_info[3] == 'E' else -decimal_degrees

    return gps_coordinates


# ------      Parsers  -------
def parse_image(file_name):
    try:
        image = Image.open(file_name)
        text = pytesseract.image_to_string(image)
        metadata = {}

        if file_name.lower().endswith(('jpeg', 'jpg')):
            exifdata = image._getexif()
            if exifdata:
                for tag_id, value in exifdata.items():
                    tag = TAGS.get(tag_id, tag_id)
                    if tag in ['GPSInfo', 'Make', 'Model', 'Software', 'DateTime']:
                        metadata[tag] = value

                if 'GPSInfo' in metadata:
                    gps_coordinates = convert_gps_coordinates(metadata['GPSInfo'])
                    metadata['GPSInfo'] = gps_coordinates

        metadata['format'] = image.format
        return {'text': text, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al leer imagen: {e}'}


def parse_pdf(file_name):
    try:

        with open(file_name, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
            pdf_text = ''

            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                pdf_text += page.extract_text()

            pdf_info = pdf_reader.metadata
            metadata = {
                'Make': pdf_info.get('/Producer'),
                'DateTime': parse_date(pdf_info.get('/CreationDate')),
                'Software': pdf_info.get('/Creator')
            }

            return {'text': pdf_text, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al leer el archivo pdf: {e}'}


def parse_dwg(file_name):
    try:
        acad = comtypes.client.GetActiveObject("AutoCAD.Application")
        doc = acad.Documents.Open(file_name)

        text_content = ""
        metadata = {
            'Software': doc.Application.ActiveDocument.SummaryInfo.Author,
            'DateTime': parse_date(doc.Application.ActiveDocument.SummaryInfo.CreationDate),
            'Make': doc.Application.ActiveDocument.Application.Version
        }

        for entity in doc.ModelSpace:
            if entity.EntityName == 'AcDbText':
                text_content += entity.TextString + "\n"

        doc.Close()
        return {'text': text_content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al leer el archivo DWG: {e}'}


def parse_bmp(file_name):

    try:
        image = Image.open(file_name)
        image = image.convert("L")  # Convertir la imagen a escala de grises (opcional)

        # Realizar OCR en la imagen BMP
        text = pytesseract.image_to_string(image)

        metadata = {
            'format': image.format,
            'mode': image.mode,
            'size': image.size
            # Puedes agregar más metadatos según necesites
        }

        return {'text': text, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo BMP: {e}'}


def parse_excel(file_name):
    try:
        wb = openpyxl.load_workbook(file_name)

        # Obtener datos de todas las hojas
        excel_data = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_without_none = [cell for cell in row if (cell is not None and str(cell)[0] != '=')]
                excel_data.append(' '.join(map(str, row_without_none)))

        # Reemplazar saltos de línea por espacios en blanco
        excel_data_joined = ' '.join(excel_data)

        # Obtener metadatos
        metadata = {
            key: value for key, value in {
                'worksheets': wb.sheetnames,
                'active_sheet': wb.active.title,
                'DateTime': parse_date(wb.properties.created) if wb.properties else None,
                'title': wb.properties.title if wb.properties else None,
            }.items() if value is not None
        }

        metadata['text'] = excel_data_joined.replace('\n', ' ')

        return {'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo Excel: {e}'}


def parse_txt(file_name):
    try:
        with open(file_name, 'r') as file:
            content = file.read()

        # Obtener metadatos
        metadata = {
            'file_name': file_name,
            'file_size': os.path.getsize(file_name),
            'file_extension': '.txt',
            # Otros metadatos según sea necesario
        }

        return {'text': content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo de texto: {e}'}


def parse_docx(file_name):
    try:
        doc = Document(file_name)

        # Obtener texto del documento
        text_content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])

        # Obtener metadatos
        core_properties = doc.core_properties
        metadata = {
            'title': core_properties.title,
            'author': core_properties.author,
            'created': parse_date(core_properties.created)
            # Agrega más metadatos según sea necesario
        }

        return {'text': text_content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el documento Word: {e}'}


def clean_text(text):
    # Eliminar caracteres no ASCII usando una expresión regular
    cleaned_text = re.sub(r'\\x[a-fA-F0-9]{2}', '', text)

    # Eliminar \r y \n
    cleaned_text = cleaned_text.replace('\r', '').replace('\n', '')

    # También puedes agregar otras manipulaciones específicas según los patrones que desees eliminar
    # cleaned_text = cleaned_text.replace('patrón', '')  # Reemplazar 'patrón' con una cadena vacía

    return cleaned_text


def parse_doc(file_name):
    try:
        # Abrir el documento .doc como texto
        with olefile.OleFileIO(file_name) as ole:
            # Extraer el texto del documento de Word y manejar errores de decodificación
            text_stream = ole.openstream('WordDocument').read()
            text = text_stream.decode('latin-1', errors='replace')

            # Encontrar palabras en español en el texto del documento
            spanish_word_pattern = r'\b(?:[^\W\d_Àÿ]|[\d])+?\b'
            found_words = find_words_in_text(text, spanish_word_pattern)

            # Eliminar palabras de una sola letra
            filtered_words = [word for word in found_words if len(word) > 1]

            # Eliminar caracteres excluidos adicionales de las palabras después de la primera
            additional_excluded_chars = 'ÕþåÀÿ'
            final_filtered_words = [filtered_words[0]] + [word.translate(str.maketrans('', '', additional_excluded_chars)) for
                                                          word in filtered_words[1:]]

            # Convertir la lista a una cadena con elementos separados por un espacio
            final_text = ' '.join(final_filtered_words)

            # Obtener metadatos
            metadata = {
                'file_path': file_name,
                # Otros metadatos según sea necesario
            }

            return {'text': final_text, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el documento .doc como texto plano: {e}'}


def parsing(file_name, extension):
    if extension.lower() in ['jpeg', 'jpg', 'png']:
        return parse_image(file_name)
    elif extension.lower() == 'pdf':
        return parse_pdf(file_name)
    elif extension.lower() == 'dwg':
        return parse_dwg(file_name)
    elif extension.lower() == 'bmp':
        return parse_bmp(file_name)
    elif extension.lower() in ['xlsx', 'xls', 'xlsm', 'xlsb', 'xltx', 'xlt']:
        return parse_excel(file_name)
    elif extension.lower() == 'txt':
        return parse_txt(file_name)
    elif extension.lower() in ['docx', 'docm', 'dotx', 'dotm']:
        return parse_docx(file_name)
    elif extension.lower() in ['doc']:
        return parse_doc(file_name)  # Manejar archivos .doc antiguos
    else:
        return {'error': 'Extensión no compatible'}


def parse_with_timeout(file_name, extension, timeout_sec):
    result = {"error": None}

    def target():
        nonlocal result
        try:
            result["data"] = parsing(file_name, extension)
        except Exception as e:
            result["error"] = str(e)

    thread = threading.Thread(target=target)
    thread.start()
    thread.join(timeout_sec)

    if thread.is_alive():
        # Si el hilo aún está vivo después del tiempo de espera, lo cancelamos
        thread.join()
        result = {"error": "Timeout excedido"}

    return result


# Test
# Reemplaza "nombre_del_archivo.doc" con la ruta de tu archivo .doc específico
file_path = "nombre_del_archivo.doc"
result = parse_doc(file_path)
print(result)
