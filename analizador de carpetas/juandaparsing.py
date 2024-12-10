from PIL import Image
import pytesseract
from PIL.ExifTags import TAGS
import PyPDF2
import comtypes.client
import openpyxl
import xlrd
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from docx import Document
import win32com.client
import os
import textract
import re
import threading
import regex
import olefile
import imageio
import subprocess
import time
import pyautogui
import keyboard
import shutil


#  ------    Tools  ------- 
def parse_date(date_string):
    try:
        if isinstance(date_string, datetime):
            # Si la entrada es del tipo datetime.datetime, devolverla directamente
            return date_string.strftime('%Y-%m-%d %H:%M:%S')

        # Eliminar 'D:' si está presente en el formato
        date_string = date_string.replace('D:', '', 1) if date_string.startswith('D:') else date_string
        
        # Convertir la cadena de fecha a un objeto datetime
        parsed_date = datetime.strptime(date_string, '%Y%m%d%H%M%S')
        
        # Formatear la fecha en un formato legible
        formatted_date = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
        
        return formatted_date
    except ValueError as e:
        return f'Error de formato: {e}'

def convert_gps_coordinates(gps_info):
    gps_coordinates = {}

    for key in gps_info:
        if key in [2, 4]:
            # Obteniendo los grados, minutos y segundos
            degrees, minutes, seconds = gps_info[key]

            # Convirtiendo a grados decimales
            decimal_degrees = degrees + (minutes / 60.0) + (seconds / 3600.0)

            # Agregando al diccionario de coordenadas GPS
            if key == 2:
                gps_coordinates['Latitude'] = decimal_degrees if gps_info[1] == 'N' else -decimal_degrees
            elif key == 4:
                gps_coordinates['Longitude'] = decimal_degrees if gps_info[3] == 'E' else -decimal_degrees

    return gps_coordinates

def remover_comas_consecutivas(cadena):
    # Utilizamos una expresión regular para encontrar todas las comas consecutivas
    # y las reemplazamos por un solo espacio
    nueva_cadena = re.sub(',+', ' ', cadena)
    return nueva_cadena

#  ------      Parsers  ------- 
def parse_image(file_name):
    try:
        if file_name.lower().endswith(('jpeg', 'jpg', 'png')):
            image = Image.open(file_name)
        elif file_name.lower().endswith('gif'):
            gif = imageio.get_reader(file_name)
            # Obtener el primer cuadro del GIF para procesarlo
            image = Image.fromarray(gif.get_data(0))

        text = pytesseract.image_to_string(image)
        metadata = {}

        if file_name.lower().endswith(('jpeg', 'jpg')):
            exifdata = image._getexif()
            if exifdata:
                for tag_id, value in exifdata.items():
                    tag = TAGS.get(tag_id, tag_id)
                    if tag in ['GPSInfo', 'Make', 'Model', 'Software', 'DateTime']:
                        metadata[tag] = value

                if 'GPSInfo' in metadata:
                    gps_coordinates = convert_gps_coordinates(metadata['GPSInfo'])
                    metadata['GPSInfo'] = gps_coordinates

        metadata['format'] = image.format
        return {'text': text, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al leer imagen: {e}'}


def parse_pdf(file_name):
    try:
    
        with open(file_name, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
            pdf_text = ''

            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                pdf_text += page.extract_text()

            pdf_info = pdf_reader.metadata
            metadata = {
                'Make': pdf_info.get('/Producer'),
                'DateTime': parse_date(pdf_info.get('/CreationDate')),
                'Software': pdf_info.get('/Creator')
            }

            return {'text': pdf_text, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al leer el archivo pdf: {e}'}

carpeta_temporal = r'C:\Script\carpeta temporal pdfs'
def limpiar_carpeta_temporal():
    try:
        # Elimina el contenido de la carpeta temporal
        for archivo in os.listdir(carpeta_temporal):
            ruta_archivo = os.path.join(carpeta_temporal, archivo)
            if os.path.isfile(ruta_archivo):
                os.remove(ruta_archivo)

        print(f"Contenido de la carpeta temporal ha sido eliminado.")
        return True

    except Exception as e:
        print(f"Error al limpiar la carpeta temporal: {e}")
        return False

def convertir_dwg_a_pdf(ruta_archivo_dwg):
    try:
        # Comando para abrir el archivo DWG con Acrobat
        subprocess.Popen(['start', '', ruta_archivo_dwg], shell=True)

        # Espera a que se abra Acrobat y el archivo (ajusta este tiempo según sea necesario)
        time.sleep(5)

        # Utiliza keyboard para activar la ventana de Acrobat
        keyboard.press_and_release('alt+tab')
        time.sleep(3)

        # Simular pulsaciones de teclas para guardar como PDF (ajusta según sea necesario)
        pyautogui.hotkey('ctrl', 's')  # Abre el cuadro de diálogo de impresión
        time.sleep(4)

        pyautogui.press('enter')  # Presiona Enter para confirmar la impresión
        time.sleep(1)
        pyautogui.press('enter')  # Presiona Enter para confirmar confirmar la impresión
        time.sleep(1)  # Espera a que se procese la impresión (ajusta según sea necesario)

        pyautogui.hotkey('alt', 'f4')  # Cierra la ventana de impresión

        # Obtener la lista de archivos en la carpeta temporal
        archivos_temporales = os.listdir(carpeta_temporal)

        # Procesar cada archivo PDF en la carpeta temporal
        resultados_parse = []
        for nombre_archivo_pdf in archivos_temporales:
            ruta_temporal_pdf = os.path.join(carpeta_temporal, nombre_archivo_pdf)

            # Llama a la función parse_pdf_from_temp_folder
            resultado_parse = parse_pdf(ruta_temporal_pdf)

            # Agrega el resultado del análisis a la lista
            resultados_parse.append(resultado_parse)

        # Devuelve la lista de resultados del análisis
        limpiar_carpeta_temporal()
        return resultados_parse

    except Exception as e:
        # Devuelve un diccionario con el error
        return {'error': f'Error al abrir el archivo DWG con Acrobat: {e}'}
    
def parse_bmp(file_name):

    try:
        image = Image.open(file_name)
        image = image.convert("L")  # Convertir la imagen a escala de grises (opcional)

        # Realizar OCR en la imagen BMP
        text = pytesseract.image_to_string(image)

        metadata = {
            'format': image.format,
            'mode': image.mode,
            'size': image.size
            # Puedes agregar más metadatos según necesites
        }

        return {'text': text, 'metadata': metadata}
    
    except Exception as e:
        return {'error': f'Error al procesar el archivo BMP: {e}'}


def parse_excel(file_name):
    try:
        if file_name.endswith('.xls'):
            # Si es un archivo .xls, usar xlrd para cargar el libro de trabajo
            xl_workbook = xlrd.open_workbook(file_name)
            sheets = xl_workbook.sheet_names()
            excel_data = []

            for sheet_name in sheets:
                sheet = xl_workbook.sheet_by_name(sheet_name)
                for row_num in range(sheet.nrows):
                    row = sheet.row_values(row_num)
                    row_without_none = [cell for cell in row if (cell is not None and str(cell)[0] != '=')]
                    excel_data.append(' '.join(map(str, row_without_none)))

        elif file_name.endswith('.xlsx'):
            # Si es un archivo .xlsx, usar openpyxl directamente
            wb = openpyxl.load_workbook(file_name)
            excel_data = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_without_none = [cell for cell in row if (cell is not None and str(cell)[0] != '=')]
                    excel_data.append(' '.join(map(str, row_without_none)))

        else:
            return {'error': 'Formato de archivo no compatible. Use archivos .xls o .xlsx.'}

        # Reemplazar saltos de línea por espacios en blanco
        excel_data_joined = ' '.join(excel_data)

        # Obtener metadatos
        metadata = {
            'worksheets': sheets if file_name.endswith('.xls') else wb.sheetnames,
            'active_sheet': sheets[0] if file_name.endswith('.xls') else wb.active.title,
            'DateTime': parse_date(wb.properties.created) if wb.properties else None,
            'title': wb.properties.title if wb.properties else None,
        }
        
        metadata['text'] = excel_data_joined.replace('\n', ' ')

        return {'text': metadata['text'], 'metadata': metadata}
    
    except Exception as e:
        return {'error': f'Error al procesar el archivo Excel: {e}'}


def parse_csv(file_name):
    try:
        with open(file_name, 'r') as file:
            content = remover_comas_consecutivas(file.read())

        # Obtener metadatos
        metadata = {
            'file_name': file_name,
            'file_size': os.path.getsize(file_name),
            'file_extension': '.csv',
            # Otros metadatos según sea necesario
        }

        return {'text': content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo de texto: {e}'}

def parse_txt(file_name, exten):
    try:
        with open(file_name, 'r') as file:
            content = file.read()

        # Obtener metadatos
        metadata = {
            'file_name': file_name,
            'file_size': os.path.getsize(file_name),
            'file_extension': exten,
            # Otros metadatos según sea necesario
        }

        return {'text': content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo de texto: {e}'}

def parse_docx(file_name):
    try:
        doc = Document(file_name)

        # Obtener texto del documento
        text_content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])

        # Obtener metadatos
        core_properties = doc.core_properties
        metadata = {
            'title': core_properties.title,
            'author': core_properties.author,
            'created': parse_date(core_properties.created)
            # Agrega más metadatos según sea necesario
        }

        return {'text': text_content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el documento Word: {e}'}

def clean_text(text):
    # Eliminar caracteres no ASCII usando una expresión regular
    cleaned_text = re.sub(r'\\x[a-fA-F0-9]{2}', '', text)

    # Eliminar \r y \n
    cleaned_text = cleaned_text.replace('\r', '').replace('\n', '')

    # También puedes agregar otras manipulaciones específicas según los patrones que desees eliminar
    # cleaned_text = cleaned_text.replace('patrón', '')  # Reemplazar 'patrón' con una cadena vacía

    return cleaned_text
    return cleaned_text
def find_words_in_text(text, pattern):
    matches = regex.findall(pattern, text)
    return matches

def parse_doc(file_name):
    try:
        # Abrir el documento .doc como texto
        with olefile.OleFileIO(file_name) as ole:
            # Extraer el texto del documento de Word y manejar errores de decodificación
            text_stream = ole.openstream('WordDocument').read()
            text = text_stream.decode('latin-1', errors='replace')

            # Encontrar palabras en español en el texto del documento
            spanish_word_pattern = r'\b(?:[^\W\d_Àÿ]|[\d])+?\b'
            found_words = find_words_in_text(text, spanish_word_pattern)

            # Eliminar palabras de una sola letra
            filtered_words = [word for word in found_words if len(word) > 1]

            # Eliminar caracteres excluidos adicionales de las palabras después de la primera
            additional_excluded_chars = 'ÕþåÀÿ'
            final_filtered_words = [filtered_words[0]] + [word.translate(str.maketrans('', '', additional_excluded_chars)) for
                                                          word in filtered_words[1:]]

            # Convertir la lista a una cadena con elementos separados por un espacio
            final_text = ' '.join(final_filtered_words)

            # Obtener metadatos
            metadata = {
                'file_path': file_name,
                # Otros metadatos según sea necesario
            }

            return {'text': final_text, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el documento .doc como texto plano: {e}'}

def parsing(file_name, extension):
    if extension.lower() in ['jpeg', 'jpg', 'png', 'gif']:
        return parse_image(file_name)
    elif extension.lower() == 'pdf':
        return parse_pdf(file_name)
    elif extension.lower() == 'dwg':
        return convertir_dwg_a_pdf(file_name)
    elif extension.lower() == 'bmp':
        return parse_bmp(file_name)
    elif extension.lower() in ['xlsx', 'xls', 'xlsm', 'xlsb', 'xltx', 'xlt']:
        return parse_excel(file_name)
    elif extension.lower() in ['txt', 'src', 'out', 'ipt', 'plt', 'bat']:
        return parse_txt(file_name)
    elif extension.lower() == 'csv':
        return parse_csv(file_name)
    elif extension.lower() in ['docx', 'docm', 'dotx', 'dotm']:
        return parse_docx(file_name)
    elif extension.lower() in ['doc']:
        return parse_doc(file_name)  # Manejar archivos .doc antiguos
    else:
        return {'error': 'Extensión no compatible'}

def parse_with_timeout(file_name, extension, timeout_sec):
    result = {"error": None}

    def target():
        nonlocal result
        try:
            result["data"] = parsing(file_name, extension)
        except Exception as e:
            result["error"] = str(e)

    thread = threading.Thread(target=target)
    thread.start()
    thread.join(timeout_sec)

    if thread.is_alive():
        # Si el hilo aún está vivo después del tiempo de espera, lo cancelamos
        thread.join()
        result["data"]["error"] = "Timeout excedido"

    return result
# Test
file_path = r'C:\Script\Son2.dwg'
file_extension = 'dwg'

result = parse_with_timeout(file_path, file_extension,30)
print(result)



