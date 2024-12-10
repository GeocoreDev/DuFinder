from PIL import Image
import pytesseract
from PIL.ExifTags import TAGS
import PyPDF2
from datetime import datetime
import comtypes.client
import openpyxl
from docx import Document
import win32com.client
import xml.etree.ElementTree as ET
import os

#  ------    Tools  ------- 
def parse_date(date_string):
    try:
        if isinstance(date_string, datetime):
            # Si la entrada es del tipo datetime.datetime, devolverla directamente
            return date_string.strftime('%Y-%m-%d %H:%M:%S')

        # Eliminar 'D:' si está presente en el formato
        date_string = date_string.replace('D:', '', 1) if date_string.startswith('D:') else date_string
        
        # Convertir la cadena de fecha a un objeto datetime
        parsed_date = datetime.strptime(date_string, '%Y%m%d%H%M%S')
        
        # Formatear la fecha en un formato legible
        formatted_date = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
        
        return formatted_date
    except ValueError as e:
        return f'Error de formato: {e}'

def convert_gps_coordinates(gps_info):
    gps_coordinates = {}

    for key in gps_info:
        if key in [2, 4]:
            # Obteniendo los grados, minutos y segundos
            degrees, minutes, seconds = gps_info[key]

            # Convirtiendo a grados decimales
            decimal_degrees = degrees + (minutes / 60.0) + (seconds / 3600.0)

            # Agregando al diccionario de coordenadas GPS
            if key == 2:
                gps_coordinates['Latitude'] = decimal_degrees if gps_info[1] == 'N' else -decimal_degrees
            elif key == 4:
                gps_coordinates['Longitude'] = decimal_degrees if gps_info[3] == 'E' else -decimal_degrees

    return gps_coordinates


#  ------      Parsers  ------- 
def parse_image(file_name):
    image = Image.open(file_name)
    text = pytesseract.image_to_string(image)
    metadata = {}

    if file_name.lower().endswith(('jpeg', 'jpg')):
        exifdata = image._getexif()
        if exifdata:
            for tag_id, value in exifdata.items():
                tag = TAGS.get(tag_id, tag_id)
                if tag in ['GPSInfo', 'Make', 'Model', 'Software', 'DateTime']:
                    metadata[tag] = value

            if 'GPSInfo' in metadata:
                gps_coordinates = convert_gps_coordinates(metadata['GPSInfo'])
                metadata['GPSInfo'] = gps_coordinates

    metadata['format'] = image.format
    return {'text': text, 'metadata': metadata}

def parse_pdf(file_name):
    with open(file_name, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        num_pages = len(pdf_reader.pages)
        pdf_text = ''

        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            pdf_text += page.extract_text()

        pdf_info = pdf_reader.metadata
        metadata = {
            'Make': pdf_info.get('/Producer'),
            'DateTime': parse_date(pdf_info.get('/CreationDate')),
            'Software': pdf_info.get('/Creator')
        }

        return {'text': pdf_text, 'metadata': metadata}

def parse_dwg(file_name):
    try:
        acad = comtypes.client.GetActiveObject("AutoCAD.Application")
        doc = acad.Documents.Open(file_name)

        text_content = ""
        metadata = {
            'Software': doc.Application.ActiveDocument.SummaryInfo.Author,
            'DateTime': parse_date(doc.Application.ActiveDocument.SummaryInfo.CreationDate),
            'Make': doc.Application.ActiveDocument.Application.Version
        }

        for entity in doc.ModelSpace:
            if entity.EntityName == 'AcDbText':
                text_content += entity.TextString + "\n"

        doc.Close()
        return {'text': text_content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al leer el archivo DWG: {e}'}

def parse_bmp(file_name):

    try:
        image = Image.open(file_name)
        image = image.convert("L")  # Convertir la imagen a escala de grises (opcional)

        # Realizar OCR en la imagen BMP
        text = pytesseract.image_to_string(image)

        metadata = {
            'format': image.format,
            'mode': image.mode,
            'size': image.size
            # Puedes agregar más metadatos según necesites
        }

        return {'text': text, 'metadata': metadata}
    
    except Exception as e:
        return {'error': f'Error al procesar el archivo BMP: {e}'}


def parse_excel(file_name):
    try:
        wb = openpyxl.load_workbook(file_name)

        # Obtener datos de todas las hojas
        excel_data = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_without_none = [cell for cell in row if (cell is not None and str(cell)[0] != '=')]
                excel_data.append(' '.join(map(str, row_without_none)))

        # Reemplazar saltos de línea por espacios en blanco
        excel_data_joined = ' '.join(excel_data)

        # Obtener metadatos
        metadata = {
            key: value for key, value in {
                'worksheets': wb.sheetnames,
                'active_sheet': wb.active.title,
                'DateTime': parse_date(wb.properties.created) if wb.properties else None,
                'title': wb.properties.title if wb.properties else None,
            }.items() if value is not None
        }
        
        metadata['text'] = excel_data_joined.replace('\n', ' ')

        return {'metadata': metadata}
    
    except Exception as e:
        return {'error': f'Error al procesar el archivo Excel: {e}'}

def parse_txt(file_name):
    try:
        with open(file_name, 'r') as file:
            content = file.read()

        # Obtener metadatos
        metadata = {
            'file_name': file_name,
            'file_size': os.path.getsize(file_name),
            'file_extension': '.txt',
            # Otros metadatos según sea necesario
        }

        return {'data': content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo de texto: {e}'}

def parse_docx(file_name):
    try:
        doc = Document(file_name)

        # Obtener texto del documento
        text_content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])

        # Obtener metadatos
        core_properties = doc.core_properties
        metadata = {
            'title': core_properties.title,
            'author': core_properties.author,
            'created': parse_date(core_properties.created)
            # Agrega más metadatos según sea necesario
        }

        return {'data': text_content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el documento Word: {e}'}

def parse_csv(file_name):
    try:
        with open(file_name, 'r') as file:
            content = file.read()

        # Obtener metadatos
        metadata = {
            'file_name': file_name,
            'file_size': os.path.getsize(file_name),
            'file_extension': '.csv',
            # Otros metadatos según sea necesario
        }

        return {'data': content, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo de texto: {e}'}


def parse_xml(file_name):
    try:
        tree = ET.parse(file_name)
        root = tree.getroot()

        # Obtener metadatos
        metadata = {
            'file_name': file_name,
            'file_size': os.path.getsize(file_name),
            'file_extension': '.xml',
            # Otros metadatos según sea necesario
        }

        # Obtener datos del XML
        xml_data = ET.tostring(root, encoding='utf-8', method='xml').decode()

        return {'data': xml_data, 'metadata': metadata}

    except Exception as e:
        return {'error': f'Error al procesar el archivo XML: {e}'}


def parsing(file_name, extension):
    if extension.lower() in ['jpeg', 'jpg', 'png']:
        return parse_image(file_name)
    elif extension.lower() == 'pdf':
        return parse_pdf(file_name)
    elif extension.lower() == 'dwg':
        return parse_dwg(file_name)
    elif extension.lower() == 'bmp':
        return parse_bmp(file_name)
    elif extension.lower() in ['xlsx', 'xls', 'xlsm', 'xlsb', 'xltx', 'xlt']:
        return parse_excel(file_name)
    elif extension.lower() == 'txt':
        return parse_txt(file_name)
    elif extension.lower() == 'csv':
        return parse_txt(file_name)
    elif extension.lower() == 'doc':
        return parse_doc(file_name)
    elif extension.lower() == 'kml':
        return parse_xml(file_name)
    elif extension.lower() in ['docx', 'docm', 'dotx', 'dotm']:
        return parse_docx(file_name)
    else:
        return {'error': 'Extensión no compatible'}

# Test
file_path = "D:/2010\GYC 0510-1603 ES DP Ruta del Sol - Tramo II EDL\Referencias\PRESENTACION RUTA DEL SOL\JCA\GPSKitData.kml"
file_extension = 'kml'
result = parsing(file_path, file_extension)
print(result)

